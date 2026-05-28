from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from docgraph_sidecar.parser.base import (
    BaseParser,
    ParsedChunk,
    ParsedDocumentElement,
    ParseContext,
    ParseResult,
    ParserError,
)


DEFAULT_PREVIEW_ROWS = 20
MAX_CELL_TEXT_LENGTH = 120
KEY_COLUMN_KEYWORDS = {
    "id",
    "name",
    "title",
    "item",
    "project",
    "customer",
    "vendor",
    "owner",
    "status",
    "date",
    "quantity",
    "amount",
    "total",
    "budget",
    "cost",
    "price",
    "编号",
    "名称",
    "项目",
    "客户",
    "供应商",
    "负责人",
    "状态",
    "日期",
    "数量",
}
AMOUNT_COLUMN_KEYWORDS = {
    "amount",
    "total",
    "budget",
    "cost",
    "price",
    "fee",
    "payment",
    "金额",
    "合计",
    "总计",
    "预算",
    "成本",
    "费用",
    "单价",
}


@dataclass(frozen=True)
class SheetProfile:
    sheet_name: str
    headers: tuple[str, ...]
    preview_rows: tuple[tuple[str, ...], ...]
    key_columns: tuple[str, ...]
    amount_columns: tuple[str, ...]
    row_count: int
    column_count: int

    def to_metadata(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "headers": list(self.headers),
            "preview_rows": [list(row) for row in self.preview_rows],
            "key_columns": list(self.key_columns),
            "amount_columns": list(self.amount_columns),
            "row_count": self.row_count,
            "column_count": self.column_count,
        }


class XlsxParser(BaseParser):
    name = "xlsx"
    supported_extensions = (".xlsx",)

    def __init__(self, *, preview_rows: int = DEFAULT_PREVIEW_ROWS) -> None:
        self.preview_rows = preview_rows

    def parse(self, context: ParseContext) -> ParseResult:
        try:
            workbook = load_workbook(context.path, read_only=True, data_only=True)
        except Exception as exc:
            raise ParserError(
                "XLSX file could not be parsed.",
                error_code="XLSX_PARSE_ERROR",
                parser_name=self.name,
                retryable=False,
                details={"path": str(context.path), "error_type": type(exc).__name__},
            ) from exc

        elements: list[ParsedDocumentElement] = []
        chunks: list[ParsedChunk] = []
        profiles: list[SheetProfile] = []
        try:
            for worksheet in workbook.worksheets:
                profile = profile_sheet(worksheet, preview_rows=self.preview_rows)
                profiles.append(profile)
                if not profile.headers and not profile.preview_rows:
                    continue
                text = sheet_profile_text(profile)
                add_sheet_block(elements, chunks, context=context, profile=profile, text=text)
        finally:
            workbook.close()

        warnings = () if elements else ("No parseable XLSX sheets found.",)
        return ParseResult(
            parser_name=self.name,
            file_id=context.file_id,
            elements=tuple(elements),
            chunks=tuple(chunks),
            warnings=warnings,
            metadata={
                "excel_profile": {
                    "sheet_count": len(profiles),
                    "sheets": [profile.to_metadata() for profile in profiles],
                }
            },
        )


def profile_sheet(worksheet: Worksheet, *, preview_rows: int = DEFAULT_PREVIEW_ROWS) -> SheetProfile:
    rows_seen = 0
    max_columns = 0
    header_values: tuple[str, ...] = ()
    preview_values: list[tuple[str, ...]] = []
    amount_scores: dict[int, int] = {}

    for raw_row in worksheet.iter_rows(values_only=True):
        row_values = tuple(normalize_cell(value) for value in raw_row)
        if not any(row_values):
            continue
        rows_seen += 1
        max_columns = max(max_columns, len(row_values))
        if not header_values:
            header_values = row_values
            continue
        if len(preview_values) < preview_rows:
            preview_values.append(row_values)
        for index, value in enumerate(row_values):
            if looks_like_currency_amount(value):
                amount_scores[index] = amount_scores.get(index, 0) + 1

    headers = trim_trailing_empty(header_values)
    key_columns = tuple(
        header
        for header in headers
        if header and any(keyword.casefold() in header.casefold() for keyword in KEY_COLUMN_KEYWORDS)
    )
    amount_columns = detect_amount_columns(headers, amount_scores)

    return SheetProfile(
        sheet_name=worksheet.title,
        headers=headers,
        preview_rows=tuple(trim_trailing_empty(row) for row in preview_values),
        key_columns=key_columns,
        amount_columns=amount_columns,
        row_count=rows_seen,
        column_count=max_columns,
    )


def detect_amount_columns(headers: tuple[str, ...], amount_scores: dict[int, int]) -> tuple[str, ...]:
    columns: list[str] = []
    for index, header in enumerate(headers):
        header_match = header and any(
            keyword.casefold() in header.casefold() for keyword in AMOUNT_COLUMN_KEYWORDS
        )
        value_match = not header and amount_scores.get(index, 0) > 0
        if header_match or value_match:
            columns.append(header or f"Column {index + 1}")
    return tuple(columns)


def sheet_profile_text(profile: SheetProfile) -> str:
    lines = [f"Sheet: {profile.sheet_name}"]
    if profile.headers:
        lines.append(f"Headers: {', '.join(profile.headers)}")
    if profile.key_columns:
        lines.append(f"Key columns: {', '.join(profile.key_columns)}")
    if profile.amount_columns:
        lines.append(f"Amount columns: {', '.join(profile.amount_columns)}")
    if profile.preview_rows:
        lines.append("Preview rows:")
        for row in profile.preview_rows:
            lines.append(" | ".join(row))
    return "\n".join(lines)


def add_sheet_block(
    elements: list[ParsedDocumentElement],
    chunks: list[ParsedChunk],
    *,
    context: ParseContext,
    profile: SheetProfile,
    text: str,
) -> None:
    index = len(elements)
    element_id = stable_parse_id(context.file_id, "element", index, text)
    chunk_id = stable_parse_id(context.file_id, "chunk", index, text)
    metadata = profile.to_metadata()
    elements.append(
        ParsedDocumentElement(
            element_id=element_id,
            file_id=context.file_id,
            element_index=index,
            element_type="sheet",
            sheet_name=profile.sheet_name,
            section_path=profile.sheet_name,
            text=text,
            metadata=metadata,
            confidence=1.0,
        )
    )
    chunks.append(
        ParsedChunk(
            chunk_id=chunk_id,
            file_id=context.file_id,
            element_id=element_id,
            chunk_index=index,
            chunk_type="sheet",
            sheet_name=profile.sheet_name,
            heading=profile.sheet_name,
            section_path=profile.sheet_name,
            text=text,
            token_count=estimate_token_count(text),
            evidence={"parser": XlsxParser.name, "preview_rows": len(profile.preview_rows)},
        )
    )


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        text = f"{value:.6f}".rstrip("0").rstrip(".")
    elif isinstance(value, Decimal):
        text = str(value.normalize())
    else:
        text = str(value)
    text = re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip()
    if len(text) > MAX_CELL_TEXT_LENGTH:
        return f"{text[:MAX_CELL_TEXT_LENGTH]}..."
    return text


def trim_trailing_empty(row: tuple[str, ...]) -> tuple[str, ...]:
    values = list(row)
    while values and values[-1] == "":
        values.pop()
    return tuple(values)


def looks_like_currency_amount(value: str) -> bool:
    if not value:
        return False
    if not any(symbol in value for symbol in ("$", "¥", "￥")):
        return False
    cleaned = value.replace(",", "").replace("$", "").replace("¥", "").replace("￥", "")
    try:
        float(cleaned)
    except ValueError:
        return False
    return True


def stable_parse_id(file_id: str, kind: str, index: int, text: str) -> str:
    digest = hashlib.sha256(f"{file_id}:{kind}:{index}:{text}".encode("utf-8")).hexdigest()
    return f"{kind}-{digest[:24]}"


def estimate_token_count(text: str) -> int:
    return len(re.findall(r"\S+", text))
