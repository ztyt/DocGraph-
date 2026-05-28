import json
import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook

from docgraph_sidecar.core.db import connect, initialize_database
from docgraph_sidecar.parser import (
    ParseContext,
    ParserError,
    XlsxParser,
    default_parser_registry,
    parse_with_error_recording,
)


class XlsxParserTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.parser = XlsxParser(preview_rows=2)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_xlsx_parser_extracts_sheet_profile_and_limited_preview(self) -> None:
        path = self.root / "budget.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Budget"
        sheet.append(["Item", "Amount", "Owner", "Notes"])
        sheet.append(["Camera", 1200.5, "North Center", "first row"])
        sheet.append(["Switch", 300, "North Center", "second row"])
        sheet.append(["Cable", 50, "Warehouse", "third row should not appear"])
        workbook.save(path)

        result = self.parser.parse(ParseContext.from_path(path, file_id="file-xlsx"))

        self.assertEqual(result.parser_name, "xlsx")
        self.assertEqual(len(result.elements), 1)
        element = result.elements[0]
        self.assertEqual(element.element_type, "sheet")
        self.assertEqual(element.sheet_name, "Budget")
        self.assertEqual(element.metadata["headers"], ["Item", "Amount", "Owner", "Notes"])
        self.assertEqual(element.metadata["preview_rows"], [
            ["Camera", "1200.5", "North Center", "first row"],
            ["Switch", "300", "North Center", "second row"],
        ])
        self.assertEqual(element.metadata["key_columns"], ["Item", "Amount", "Owner"])
        self.assertEqual(element.metadata["amount_columns"], ["Amount"])
        self.assertIn("Headers: Item, Amount, Owner, Notes", result.chunks[0].text)
        self.assertIn("Camera | 1200.5 | North Center | first row", result.chunks[0].text)
        self.assertNotIn("third row should not appear", result.chunks[0].text)
        self.assertEqual(result.metadata["excel_profile"]["sheet_count"], 1)

    def test_multiple_sheets_create_multiple_elements(self) -> None:
        path = self.root / "multi.xlsx"
        workbook = Workbook()
        first = workbook.active
        first.title = "Summary"
        first.append(["Project", "Budget"])
        first.append(["Alpha", 5000])
        second = workbook.create_sheet("Inventory")
        second.append(["Item", "Quantity"])
        second.append(["Camera", 12])
        workbook.save(path)

        result = self.parser.parse(ParseContext.from_path(path, file_id="file-multi"))

        self.assertEqual([element.sheet_name for element in result.elements], ["Summary", "Inventory"])
        self.assertEqual(result.elements[0].metadata["amount_columns"], ["Budget"])
        self.assertEqual(result.elements[1].metadata["key_columns"], ["Item", "Quantity"])

    def test_basic_fixture_xlsx_can_be_parsed(self) -> None:
        path = Path("fixtures/basic_docs/alpha_budget.xlsx")

        result = self.parser.parse(ParseContext.from_path(path, file_id="file-fixture"))

        self.assertEqual(result.elements[0].sheet_name, "Budget")
        self.assertEqual(result.elements[0].metadata["headers"], ["Item", "Quantity"])
        self.assertEqual(result.elements[0].metadata["amount_columns"], [])
        self.assertIn("Alpha Project camera", result.chunks[0].text)

    def test_default_registry_selects_xlsx_parser(self) -> None:
        path = self.root / "simple.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Total"])
        sheet.append(["Alpha", 12])
        workbook.save(path)

        result = default_parser_registry().parse_path(path, file_id="file-registry")

        self.assertEqual(result.parser_name, "xlsx")
        self.assertEqual(result.elements[0].metadata["amount_columns"], ["Total"])

    def test_bad_xlsx_records_parse_error(self) -> None:
        data_dir = self.root / "data"
        bad_path = self.root / "bad.xlsx"
        bad_path.write_bytes(b"not a workbook")
        initialize_database(data_dir=data_dir)
        self._insert_file(data_dir, "file-bad", str(bad_path))
        context = ParseContext.from_path(bad_path, file_id="file-bad")

        with self.assertRaises(ParserError):
            parse_with_error_recording(default_parser_registry(), context, data_dir=data_dir)

        connection = connect(data_dir=data_dir)
        try:
            row = connection.execute(
                """
                SELECT error_code, error_message, parser_name, retryable, details_json
                FROM parse_errors
                WHERE file_id = 'file-bad'
                """
            ).fetchone()
        finally:
            connection.close()

        self.assertIsNotNone(row)
        self.assertEqual(row["error_code"], "XLSX_PARSE_ERROR")
        self.assertEqual(row["parser_name"], "xlsx")
        self.assertEqual(row["retryable"], 0)
        self.assertEqual(row["error_message"], "XLSX file could not be parsed.")
        self.assertEqual(json.loads(row["details_json"])["path"], str(bad_path))

    def _insert_file(self, data_dir: Path, file_id: str, path: str) -> None:
        connection = connect(data_dir=data_dir)
        try:
            connection.execute(
                """
                INSERT INTO files (
                  file_id,
                  path,
                  normalized_path,
                  filename,
                  extension,
                  source_type,
                  file_status,
                  parse_status,
                  deleted_flag
                )
                VALUES (?, ?, ?, ?, '.xlsx', 'office', 'discovered', 'pending', 0)
                """,
                (file_id, path, path.casefold(), Path(path).name),
            )
            connection.commit()
        finally:
            connection.close()


if __name__ == "__main__":
    unittest.main()
